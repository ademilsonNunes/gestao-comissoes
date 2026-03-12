from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import hashlib
import unicodedata

from openpyxl import load_workbook

from ..config import ARTFATOS_DIR
from ..database.database import get_conn, init_schema
from ..database.models import (
    criar_ou_substituir_apuracao_periodo,
    inserir_lancamentos,
    listar_codvend_distintos,
    upsert_representante,
)


def _normalize_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return text.lower().strip()


def _arquivo_hash_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _check_hash_conflict(file_hash: str, mes: int, ano: int) -> None:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT mes, ano FROM importacoes WHERE arquivo_hash=?", (file_hash,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return
    mes_old, ano_old = int(row[0] or 0), int(row[1] or 0)
    if mes_old != mes or ano_old != ano:
        raise ValueError(
            f"arquivo_ja_importado_em_outro_periodo: arquivo ja importado em {mes_old:02d}/{ano_old}, "
            f"novo periodo solicitado {mes:02d}/{ano}"
        )


def _registrar_importacao(path: Path, file_hash: str, mes: int, ano: int, dbc: int, devolucoes: int, apuracao_id: int) -> None:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO importacoes(apuracao_id, arquivo_nome, arquivo_hash, mes, ano, dbc, devolucoes, data_importacao)
        VALUES(?,?,?,?,?,?,?,?)
        ON CONFLICT(arquivo_hash) DO UPDATE SET
            apuracao_id=excluded.apuracao_id,
            arquivo_nome=excluded.arquivo_nome,
            mes=excluded.mes,
            ano=excluded.ano,
            dbc=excluded.dbc,
            devolucoes=excluded.devolucoes,
            data_importacao=excluded.data_importacao
        """,
        (int(apuracao_id or 0) or None, path.name, file_hash, mes, ano, dbc, devolucoes, datetime.utcnow().isoformat()),
    )
    conn.commit()
    conn.close()


def _replace_period_data(mes: int, ano: int) -> None:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM apuracoes WHERE mes=? AND ano=?", (mes, ano))
    row = cur.fetchone()
    if row:
        apuracao_id = int(row[0] or 0)
        cur.execute("DELETE FROM ajustes WHERE lancamento_id IN (SELECT id FROM lancamentos WHERE apuracao_id=?)", (apuracao_id,))
        cur.execute("DELETE FROM importacoes WHERE apuracao_id=?", (apuracao_id,))
        cur.execute("DELETE FROM comissoes WHERE apuracao_id=?", (apuracao_id,))
        cur.execute("DELETE FROM lancamentos WHERE apuracao_id=?", (apuracao_id,))
        cur.execute("DELETE FROM apuracoes WHERE id=?", (apuracao_id,))
    cur.execute("DELETE FROM lancamentos WHERE mes=? AND ano=?", (mes, ano))
    cur.execute("DELETE FROM comissoes WHERE mes=? AND ano=?", (mes, ano))
    cur.execute("DELETE FROM importacoes WHERE mes=? AND ano=?", (mes, ano))
    conn.commit()
    conn.close()


def importar_representantes_base() -> int:
    init_schema()
    arquivo = ARTFATOS_DIR / "BASE ENVIO COMISSAO JANEIRO.xlsm"
    wb = load_workbook(filename=str(arquivo), data_only=True)
    ws = wb["Planilha1"] if "Planilha1" in wb.sheetnames else wb.active
    count = 0
    mapa_codvend_nome: Dict[str, str] = {}
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT codvend, vend FROM lancamentos WHERE COALESCE(codvend,'')<>''")
        for cod, nome_vend in cur.fetchall():
            key = _normalize_name(nome_vend or "")
            if key and key not in mapa_codvend_nome:
                mapa_codvend_nome[key] = str(cod)
        conn.close()
    except Exception:
        pass

    hdr_row_idx = 1
    for r in range(1, 10):
        vals = [str(c or "").strip().lower() for c in next(ws.iter_rows(min_row=r, max_row=r, values_only=True))]
        if any(v for v in vals):
            hdr_row_idx = r
            break
    header = [str(c or "").strip().lower() for c in next(ws.iter_rows(min_row=hdr_row_idx, max_row=hdr_row_idx, values_only=True))]
    idx = {h: i for i, h in enumerate(header)}

    def col(*names):
        for n in names:
            if n in idx:
                return idx[n]
        return None

    i_cod = col("codvend", "codigo", "cod", "cÃ³digo")
    i_email = col("email", "e-mail")
    i_nome = col("nome", "name")
    i_corpo = col("corpo_email", "corpo do email", "corpo do e-mail", "mensagem", "corpo")
    start_row = hdr_row_idx + 1
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        codvend = str((row[i_cod] if i_cod is not None else row[0]) or "").strip()
        email = str((row[i_email] if i_email is not None else row[1]) or "").strip()
        nome = str((row[i_nome] if i_nome is not None else row[2]) or "").strip()
        corpo = str((row[i_corpo] if i_corpo is not None else row[3]) or "").strip()
        if not codvend and nome:
            nome_key = _normalize_name(nome)
            codvend = mapa_codvend_nome.get(nome_key, "")
            if not codvend:
                for k, v in mapa_codvend_nome.items():
                    if nome_key and (nome_key in k or k in nome_key):
                        codvend = v
                        break
        if codvend:
            upsert_representante(codvend, nome, email, corpo)
            count += 1
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("UPDATE representantes SET ativo=0 WHERE COALESCE(codvend,'')=''")
        conn.commit()
        conn.close()
    except Exception:
        pass
    if count == 0:
        for cod in listar_codvend_distintos():
            upsert_representante(cod, "", "", "")
            count += 1
    return count


def _sheet_to_dicts(ws) -> List[Dict[str, Any]]:
    headers = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    registros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for i, h in enumerate(headers):
            d[h] = row[i]
        registros.append(d)
    return registros


def _parse_date(val: Any) -> Optional[datetime]:
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
    return None


def _to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        if isinstance(v, str) and not v.strip():
            return None
        return int(float(v))
    except Exception:
        return None


def _infer_mes_ano(registros: List[Dict[str, Any]]) -> Tuple[int, int]:
    meses_col = []
    anos_col = []
    for r in registros[:2000]:
        m = _to_int(r.get("MES"))
        a = _to_int(r.get("ANO"))
        if m and 1 <= m <= 12:
            meses_col.append(m)
        if a and 2000 <= a <= 2100:
            anos_col.append(a)

    if meses_col and anos_col:
        mes = max(set(meses_col), key=meses_col.count)
        ano = max(set(anos_col), key=anos_col.count)
        return mes, ano

    meses_dt = []
    anos_dt = []
    for r in registros[:2000]:
        for k in ("DTBAIXA", "DTEMISSAO", "VENCTO"):
            dt = _parse_date(r.get(k))
            if dt:
                meses_dt.append(dt.month)
                anos_dt.append(dt.year)
                break

    if meses_dt and anos_dt:
        mes = max(set(meses_dt), key=meses_dt.count)
        ano = max(set(anos_dt), key=anos_dt.count)
        confianca = meses_dt.count(mes) / max(1, len(meses_dt))
        if confianca < 0.55:
            raise ValueError("periodo_ambiguo: informe mes/ano manualmente no upload")
        return mes, ano

    raise ValueError("periodo_nao_identificado: informe mes/ano manualmente no upload")


def _escolher_abas_vendas(path: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = load_workbook(filename=str(path), data_only=True)
    db_name = None
    dv_name = None
    for n in wb.sheetnames:
        nl = _normalize_name(n)
        if "dbcomiss" in nl:
            db_name = n
        if "devolu" in nl:
            dv_name = n
    db = wb[db_name] if db_name else wb[wb.sheetnames[0]]
    dv = wb[dv_name] if dv_name else wb[wb.sheetnames[-1]]
    return _sheet_to_dicts(db), _sheet_to_dicts(dv)


def importar_vendas() -> Dict[str, int]:
    init_schema()
    arquivo = ARTFATOS_DIR / "ComissÃ£o MarÃ§o 2026.xlsx"
    if not arquivo.exists():
        candidatos = sorted(ARTFATOS_DIR.glob("ComissÃ£o*.xlsx"), reverse=True)
        if not candidatos:
            raise FileNotFoundError("arquivo_padrao_nao_encontrado")
        arquivo = candidatos[0]
    return importar_vendas_arquivo(arquivo, mes_override=3, ano_override=2026)


def importar_vendas_arquivo(path: Path, mes_override: Optional[int] = None, ano_override: Optional[int] = None) -> Dict[str, int]:
    init_schema()
    registros_db, registros_dv = _escolher_abas_vendas(path)
    mes_inf, ano_inf = _infer_mes_ano(registros_db or registros_dv)
    mes = int(mes_override) if mes_override else mes_inf
    ano = int(ano_override) if ano_override else ano_inf
    if not (1 <= mes <= 12):
        raise ValueError("mes_invalido")
    if not (2000 <= ano <= 2100):
        raise ValueError("ano_invalido")

    file_hash = _arquivo_hash_sha256(path)
    _check_hash_conflict(file_hash, mes, ano)
    _replace_period_data(mes, ano)
    apuracao_id = criar_ou_substituir_apuracao_periodo(
        mes=mes,
        ano=ano,
        arquivo_nome=path.name,
        arquivo_hash=file_hash,
        data_importacao=datetime.utcnow().isoformat(),
    )

    for r in registros_db:
        r["MES"] = mes
        r["ANO"] = ano
    for r in registros_dv:
        r["MES"] = mes
        r["ANO"] = ano

    inserir_lancamentos(registros_db, "DB", apuracao_id=apuracao_id)
    inserir_lancamentos(registros_dv, "DS", apuracao_id=apuracao_id)
    _registrar_importacao(path, file_hash, mes, ano, len(registros_db), len(registros_dv), apuracao_id=apuracao_id)
    return {"dbc": len(registros_db), "devolucoes": len(registros_dv), "mes": mes, "ano": ano, "apuracao_id": apuracao_id}


def _add_months(d: date, months: int) -> date:
    y = d.year + ((d.month - 1 + months) // 12)
    m = ((d.month - 1 + months) % 12) + 1
    return date(y, m, 1)


def _periodo_query(mes: int, ano: int) -> Tuple[date, date, date, date]:
    dtbaixa_ini = date(int(ano), int(mes), 1)
    dtbaixa_fim = _add_months(dtbaixa_ini, 1) - timedelta(days=1)
    dtemissao_ini = date.today() - timedelta(days=365)
    dtemissao_fim = date.today()
    return dtemissao_ini, dtemissao_fim, dtbaixa_ini, dtbaixa_fim


def _to_yyyymmdd(d: date) -> str:
    return d.strftime("%Y%m%d")


def _to_iso_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return s


def _norm_row_sql(row: Dict[str, Any], mes: int, ano: int) -> Dict[str, Any]:
    out = {str(k or "").upper(): v for k, v in row.items()}
    out["DTEMISSAO"] = _to_iso_date(out.get("DTEMISSAO"))
    out["VENCTO"] = _to_iso_date(out.get("VENCTO"))
    out["DTBAIXA"] = _to_iso_date(out.get("DTBAIXA"))
    out["MES"] = int(mes)
    out["ANO"] = int(ano)
    return out


def _query_comissoes_unificada(dtini: str, dtfim: str, dtbaixa_ini: str, dtbaixa_fim: str) -> str:
    return f"""
DECLARE @DTINI AS VARCHAR(8) = '{dtini}';
DECLARE @DTFIN AS VARCHAR(8) = '{dtfim}';
DECLARE @DTBXI AS VARCHAR(8) = '{dtbaixa_ini}';
DECLARE @DTBXF AS VARCHAR(8) = '{dtbaixa_fim}';

;WITH
CTE_VENDEDORES AS (
    SELECT V.A3_COD AS CODVEND, V.A3_NOME AS VEND, V.A3_COMIS AS COMIS_VEND, V.A3_SUPER AS COD_SUPER, S.A3_NOME AS NOME_SUPER
    FROM SA3010 V
    LEFT JOIN SA3010 S ON S.A3_COD = V.A3_SUPER AND S.D_E_L_E_T_ = ''
    WHERE V.D_E_L_E_T_ = ''
),
CTE_NF_JMT AS (
    SELECT DISTINCT E1_NUM AS NF, E1_SERIE AS SERIE
    FROM SE1020 INNER JOIN CTE_VENDEDORES V ON V.CODVEND = E1_VEND1
    WHERE D_E_L_E_T_ = '' AND E1_EMISSAO BETWEEN @DTINI AND @DTFIN AND E1_ZZPGCOM IN ('','N')
      AND E1_CLIENTE NOT IN ('002268','015699') AND E1_VEND1 <> '000707'
      AND E1_BAIXA BETWEEN @DTBXI AND @DTBXF AND E1_SALDO = 0
),
CTE_NF_3F AS (
    SELECT DISTINCT E1_NUM AS NF, E1_SERIE AS SERIE
    FROM SE1040 INNER JOIN CTE_VENDEDORES V ON V.CODVEND = E1_VEND1
    WHERE D_E_L_E_T_ = '' AND E1_EMISSAO BETWEEN @DTINI AND @DTFIN AND E1_ZZPGCOM IN ('','N')
      AND E1_CLIENTE NOT IN ('002268','015699') AND E1_VEND1 <> '000707'
      AND E1_BAIXA BETWEEN @DTBXI AND @DTBXF AND E1_SALDO = 0
),
CTE_NF_COS AS (
    SELECT DISTINCT E1_NUM AS NF, E1_SERIE AS SERIE
    FROM SE1050 INNER JOIN CTE_VENDEDORES V ON V.CODVEND = E1_VEND1
    WHERE D_E_L_E_T_ = '' AND E1_EMISSAO BETWEEN @DTINI AND @DTFIN AND E1_ZZPGCOM IN ('','N')
      AND E1_CLIENTE NOT IN ('002268','015699') AND E1_VEND1 <> '000707'
      AND E1_BAIXA BETWEEN @DTBXI AND @DTBXF AND E1_SALDO = 0
),
CTE_NF_SOB AS (
    SELECT DISTINCT E1_NUM AS NF, E1_SERIE AS SERIE
    FROM SE1010 INNER JOIN CTE_VENDEDORES V ON V.CODVEND = E1_VEND1
    WHERE D_E_L_E_T_ = '' AND E1_EMISSAO BETWEEN @DTINI AND @DTFIN AND E1_ZZPGCOM IN ('','N')
      AND E1_CLIENTE NOT IN ('002268','015699') AND E1_VEND1 <> '000707'
      AND E1_BAIXA BETWEEN @DTBXI AND @DTBXF AND E1_SALDO = 0
),
CTE_DATAS_JMT AS (SELECT E1_NUM,E1_SERIE,MAX(E1_VENCREA) AS VENCTO,MAX(E1_BAIXA) AS DTBAIXA FROM SE1020 WHERE D_E_L_E_T_='' GROUP BY E1_NUM,E1_SERIE),
CTE_DATAS_3F  AS (SELECT E1_NUM,E1_SERIE,MAX(E1_VENCREA) AS VENCTO,MAX(E1_BAIXA) AS DTBAIXA FROM SE1040 WHERE D_E_L_E_T_='' GROUP BY E1_NUM,E1_SERIE),
CTE_DATAS_COS AS (SELECT E1_NUM,E1_SERIE,MAX(E1_VENCREA) AS VENCTO,MAX(E1_BAIXA) AS DTBAIXA FROM SE1050 WHERE D_E_L_E_T_='' GROUP BY E1_NUM,E1_SERIE),
CTE_DATAS_SOB AS (SELECT E1_NUM,E1_SERIE,MAX(E1_VENCREA) AS VENCTO,MAX(E1_BAIXA) AS DTBAIXA FROM SE1010 WHERE D_E_L_E_T_='' GROUP BY E1_NUM,E1_SERIE),
CTE_ITENS_JMT AS (
    SELECT 'JMT' AS EMP,'VJ' AS TP,V.CODVEND,V.COD_SUPER AS SUPER,V.NOME_SUPER,V.VEND,SD2.D2_DOC AS NF,SD2.D2_PEDIDO AS PEDIDO,SD2.D2_ITEM AS ITEM,SD2.D2_COD AS CODPROD,SB1.B1_DESC AS PRODUTO,
           CONVERT(DATE,CAST(SD2.D2_EMISSAO AS DATE),103) AS DTEMISSAO,CONVERT(DATE,CAST(DT.VENCTO AS DATE),103) AS VENCTO,CONVERT(DATE,CAST(DT.DTBAIXA AS DATE),103) AS DTBAIXA,
           SD2.D2_CLIENTE AS CODCLIENTE,SA1.A1_ZZREDE AS REDE,SA1.A1_EST AS UF,SA1.A1_NOME AS CLIENTE,
           CASE SA1.A1_CODSEG WHEN '004000' THEN 'D' WHEN '001000' THEN 'V' WHEN '002000' THEN 'A' ELSE 'V' END AS CLASS_CLI,
           SD2.D2_TOTAL AS VLRBRUTO,(SD2.D2_TOTAL-((SB1.B1_FTCOM*SD2.D2_TOTAL)/100)-((SD2.D2_TOTAL*SA1.A1_ZZCONTR)/100)) AS VLRLIQ,
           SA1.A1_COMIS AS COMIS_CLI,V.COMIS_VEND,
           SB1.B1_COMIS AS COMIS_PROD,
           ((SD2.D2_TOTAL-((SB1.B1_FTCOM*SD2.D2_TOTAL)/100)-((SD2.D2_TOTAL*SA1.A1_ZZCONTR)/100))
             * SB1.B1_COMIS / 100) AS TCOMISPROD
    FROM SD2020 SD2
    INNER JOIN CTE_NF_JMT NF ON NF.NF=SD2.D2_DOC AND NF.SERIE=SD2.D2_SERIE
    INNER JOIN CTE_DATAS_JMT DT ON DT.E1_NUM=SD2.D2_DOC AND DT.E1_SERIE=SD2.D2_SERIE
    INNER JOIN SB1020 SB1 ON SB1.B1_COD=SD2.D2_COD AND SB1.D_E_L_E_T_=''
    INNER JOIN SA1020 SA1 ON SA1.A1_COD=SD2.D2_CLIENTE AND SA1.D_E_L_E_T_=''
    INNER JOIN SF2020 SF2 ON SF2.F2_DOC=SD2.D2_DOC AND SF2.F2_SERIE=SD2.D2_SERIE AND SF2.D_E_L_E_T_=''
    INNER JOIN CTE_VENDEDORES V ON V.CODVEND=SF2.F2_VEND1
    WHERE SD2.D_E_L_E_T_='' AND SD2.D2_EMISSAO BETWEEN @DTINI AND @DTFIN
),
CTE_ITENS_3F AS (
    SELECT '3F' AS EMP,'V3' AS TP,V.CODVEND,V.COD_SUPER AS SUPER,V.NOME_SUPER,V.VEND,SD2.D2_DOC AS NF,SD2.D2_PEDIDO AS PEDIDO,SD2.D2_ITEM AS ITEM,SD2.D2_COD AS CODPROD,SB1.B1_DESC AS PRODUTO,
           CONVERT(DATE,CAST(SD2.D2_EMISSAO AS DATE),103) AS DTEMISSAO,CONVERT(DATE,CAST(DT.VENCTO AS DATE),103) AS VENCTO,CONVERT(DATE,CAST(DT.DTBAIXA AS DATE),103) AS DTBAIXA,
           SD2.D2_CLIENTE AS CODCLIENTE,SA1.A1_ZZREDE AS REDE,SA1.A1_EST AS UF,SA1.A1_NOME AS CLIENTE,
           CASE SA1.A1_CODSEG WHEN '004000' THEN 'D' WHEN '001000' THEN 'V' WHEN '002000' THEN 'A' ELSE 'V' END AS CLASS_CLI,
           SD2.D2_TOTAL AS VLRBRUTO,(SD2.D2_TOTAL-((SB1.B1_FTCOM*SD2.D2_TOTAL)/100)-((SD2.D2_TOTAL*SA1.A1_ZZCONTR)/100)) AS VLRLIQ,
           SA1.A1_COMIS AS COMIS_CLI,V.COMIS_VEND,
           SB1.B1_COMIS AS COMIS_PROD,
           ((SD2.D2_TOTAL-((SB1.B1_FTCOM*SD2.D2_TOTAL)/100)-((SD2.D2_TOTAL*SA1.A1_ZZCONTR)/100))
             * SB1.B1_COMIS / 100) AS TCOMISPROD
    FROM SD2040 SD2
    INNER JOIN CTE_NF_3F NF ON NF.NF=SD2.D2_DOC AND NF.SERIE=SD2.D2_SERIE
    INNER JOIN CTE_DATAS_3F DT ON DT.E1_NUM=SD2.D2_DOC AND DT.E1_SERIE=SD2.D2_SERIE
    INNER JOIN SB1040 SB1 ON SB1.B1_COD=SD2.D2_COD AND SB1.D_E_L_E_T_=''
    INNER JOIN SA1040 SA1 ON SA1.A1_COD=SD2.D2_CLIENTE AND SA1.D_E_L_E_T_=''
    INNER JOIN SF2040 SF2 ON SF2.F2_DOC=SD2.D2_DOC AND SF2.F2_SERIE=SD2.D2_SERIE AND SF2.D_E_L_E_T_=''
    INNER JOIN CTE_VENDEDORES V ON V.CODVEND=SF2.F2_VEND1
    WHERE SD2.D_E_L_E_T_='' AND SD2.D2_EMISSAO BETWEEN @DTINI AND @DTFIN
),
CTE_ITENS_COS AS (
    SELECT 'COSMETICOS' AS EMP,'VC' AS TP,V.CODVEND,V.COD_SUPER AS SUPER,V.NOME_SUPER,V.VEND,SD2.D2_DOC AS NF,SD2.D2_PEDIDO AS PEDIDO,SD2.D2_ITEM AS ITEM,SD2.D2_COD AS CODPROD,SB1.B1_DESC AS PRODUTO,
           CONVERT(DATE,CAST(SD2.D2_EMISSAO AS DATE),103) AS DTEMISSAO,CONVERT(DATE,CAST(DT.VENCTO AS DATE),103) AS VENCTO,CONVERT(DATE,CAST(DT.DTBAIXA AS DATE),103) AS DTBAIXA,
           SD2.D2_CLIENTE AS CODCLIENTE,SA1.A1_ZZREDE AS REDE,SA1.A1_EST AS UF,SA1.A1_NOME AS CLIENTE,
           CASE SA1.A1_CODSEG WHEN '004000' THEN 'D' WHEN '001000' THEN 'V' WHEN '002000' THEN 'A' ELSE 'V' END AS CLASS_CLI,
           SD2.D2_TOTAL AS VLRBRUTO,(SD2.D2_TOTAL-((SB1.B1_FTCOM*SD2.D2_TOTAL)/100)-((SD2.D2_TOTAL*SA1.A1_ZZCONTR)/100)) AS VLRLIQ,
           SA1.A1_COMIS AS COMIS_CLI,V.COMIS_VEND,
           SB1.B1_COMIS AS COMIS_PROD,
           ((SD2.D2_TOTAL-((SB1.B1_FTCOM*SD2.D2_TOTAL)/100)-((SD2.D2_TOTAL*SA1.A1_ZZCONTR)/100))
             * SB1.B1_COMIS / 100) AS TCOMISPROD
    FROM SD2050 SD2
    INNER JOIN CTE_NF_COS NF ON NF.NF=SD2.D2_DOC AND NF.SERIE=SD2.D2_SERIE
    INNER JOIN CTE_DATAS_COS DT ON DT.E1_NUM=SD2.D2_DOC AND DT.E1_SERIE=SD2.D2_SERIE
    INNER JOIN SB1050 SB1 ON SB1.B1_COD=SD2.D2_COD AND SB1.D_E_L_E_T_=''
    INNER JOIN SA1050 SA1 ON SA1.A1_COD=SD2.D2_CLIENTE AND SA1.D_E_L_E_T_=''
    INNER JOIN SF2050 SF2 ON SF2.F2_DOC=SD2.D2_DOC AND SF2.F2_SERIE=SD2.D2_SERIE AND SF2.D_E_L_E_T_=''
    INNER JOIN CTE_VENDEDORES V ON V.CODVEND=SF2.F2_VEND1
    WHERE SD2.D_E_L_E_T_='' AND SD2.D2_EMISSAO BETWEEN @DTINI AND @DTFIN
),
CTE_ITENS_SOB AS (
    SELECT 'SOBEL' AS EMP,'VS' AS TP,V.CODVEND,V.COD_SUPER AS SUPER,V.NOME_SUPER,V.VEND,SD2.D2_DOC AS NF,SD2.D2_PEDIDO AS PEDIDO,SD2.D2_ITEM AS ITEM,SD2.D2_COD AS CODPROD,SB1.B1_DESC AS PRODUTO,
           CONVERT(DATE,CAST(SD2.D2_EMISSAO AS DATE),103) AS DTEMISSAO,CONVERT(DATE,CAST(DT.VENCTO AS DATE),103) AS VENCTO,CONVERT(DATE,CAST(DT.DTBAIXA AS DATE),103) AS DTBAIXA,
           SD2.D2_CLIENTE AS CODCLIENTE,SA1.A1_ZZREDE AS REDE,SA1.A1_EST AS UF,SA1.A1_NOME AS CLIENTE,
           CASE SA1.A1_CODSEG WHEN '004000' THEN 'D' WHEN '001000' THEN 'V' WHEN '002000' THEN 'A' ELSE 'V' END AS CLASS_CLI,
           SD2.D2_TOTAL AS VLRBRUTO,(SD2.D2_TOTAL-((SB1.B1_FTCOM*SD2.D2_TOTAL)/100)-((SD2.D2_TOTAL*SA1.A1_ZZCONTR)/100)) AS VLRLIQ,
           SA1.A1_COMIS AS COMIS_CLI,V.COMIS_VEND,
           SB1.B1_COMIS AS COMIS_PROD,
           ((SD2.D2_TOTAL-((SB1.B1_FTCOM*SD2.D2_TOTAL)/100)-((SD2.D2_TOTAL*SA1.A1_ZZCONTR)/100))
             * SB1.B1_COMIS / 100) AS TCOMISPROD
    FROM SD2010 SD2
    INNER JOIN CTE_NF_SOB NF ON NF.NF=SD2.D2_DOC AND NF.SERIE=SD2.D2_SERIE
    INNER JOIN CTE_DATAS_SOB DT ON DT.E1_NUM=SD2.D2_DOC AND DT.E1_SERIE=SD2.D2_SERIE
    INNER JOIN SB1010 SB1 ON SB1.B1_COD=SD2.D2_COD AND SB1.D_E_L_E_T_=''
    INNER JOIN SA1010 SA1 ON SA1.A1_COD=SD2.D2_CLIENTE AND SA1.D_E_L_E_T_=''
    INNER JOIN SF2010 SF2 ON SF2.F2_DOC=SD2.D2_DOC AND SF2.F2_SERIE=SD2.D2_SERIE AND SF2.D_E_L_E_T_=''
    INNER JOIN CTE_VENDEDORES V ON V.CODVEND=SF2.F2_VEND1
    WHERE SD2.D_E_L_E_T_='' AND SD2.D2_EMISSAO BETWEEN @DTINI AND @DTFIN
)
SELECT DISTINCT EMP,TP,CODVEND,SUPER,NOME_SUPER AS SUPER_NOME,VEND,NF,PEDIDO,ITEM,CODPROD,PRODUTO,DTEMISSAO,VENCTO,DTBAIXA,CODCLIENTE,REDE,UF,CLIENTE,CLASS_CLI,VLRBRUTO,VLRLIQ,COMIS_CLI,COMIS_VEND,COMIS_PROD,TCOMISPROD,MONTH(DTBAIXA) AS MES,YEAR(DTBAIXA) AS ANO
FROM (
    SELECT * FROM CTE_ITENS_JMT
    UNION ALL SELECT * FROM CTE_ITENS_3F
    UNION ALL SELECT * FROM CTE_ITENS_COS
    UNION ALL SELECT * FROM CTE_ITENS_SOB
) DBCOMISS
ORDER BY EMP,DTBAIXA,CODVEND,NF,ITEM;

;WITH
CTE_VENDEDORES AS (
    SELECT V.A3_COD AS CODVEND, V.A3_NOME AS VEND, V.A3_COMIS AS COMIS_VEND, V.A3_SUPER AS COD_SUPER, S.A3_NOME AS NOME_SUPER
    FROM SA3010 V
    LEFT JOIN SA3010 S ON S.A3_COD = V.A3_SUPER AND S.D_E_L_E_T_ = ''
    WHERE V.D_E_L_E_T_ = ''
),
CTE_DEV_SOB AS (
    SELECT 'SOBEL' AS EMP,'DS' AS TP,V.CODVEND,V.COD_SUPER AS SUPER,V.NOME_SUPER,V.VEND,SD1.D1_DOC AS NF,SD2.D2_PEDIDO AS PEDIDO,SD1.D1_ITEM AS ITEM,SD2.D2_COD AS CODPROD,SB1.B1_DESC AS PRODUTO,
           CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS DTEMISSAO,CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS VENCTO,CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS DTBAIXA,
           SA1.A1_COD AS CODCLIENTE,SA1.A1_ZZREDE AS REDE,SA1.A1_EST AS UF,SA1.A1_NOME AS CLIENTE,
           CASE SA1.A1_CODSEG WHEN '004000' THEN 'D' WHEN '001000' THEN 'V' WHEN '002000' THEN 'A' ELSE 'V' END AS CLASS_CLI,
           SD1.D1_TOTAL * -1 AS VLRBRUTO,SD1.D1_TOTAL * -1 AS VLRLIQ,SA1.A1_COMIS AS COMIS_CLI,V.COMIS_VEND,
           SB1.B1_COMIS AS COMIS_PROD,
           ((SD1.D1_TOTAL * SB1.B1_COMIS) / 100) * -1 AS TCOMISPROD
    FROM SD1010 SD1
    INNER JOIN SD2010 SD2 ON SD2.D2_DOC=SD1.D1_NFORI AND SD2.D2_SERIE=SD1.D1_SERIORI AND SD2.D2_ITEM=SD1.D1_ITEMORI AND SD2.D2_CLIENTE=SD1.D1_FORNECE AND SD2.D2_LOJA=SD1.D1_LOJA AND SD2.D_E_L_E_T_=''
    INNER JOIN SB1010 SB1 ON SB1.B1_COD=SD1.D1_COD AND SB1.D_E_L_E_T_=''
    INNER JOIN SA1010 SA1 ON SA1.A1_COD=SD1.D1_FORNECE AND SA1.A1_LOJA=SD1.D1_LOJA AND SA1.D_E_L_E_T_=''
    INNER JOIN SF2010 SF2 ON SF2.F2_DOC=SD2.D2_DOC AND SF2.F2_SERIE=SD2.D2_SERIE AND SF2.D_E_L_E_T_=''
    INNER JOIN CTE_VENDEDORES V ON V.CODVEND=SF2.F2_VEND1
    WHERE SD1.D_E_L_E_T_='' AND SD1.D1_TIPO='D' AND SUBSTRING(SD1.D1_CF,2,1)<>'9'
      AND CONVERT(VARCHAR(8), SD1.D1_DTDIGIT, 112) BETWEEN @DTBXI AND @DTBXF
      AND SD1.D1_DOC NOT IN ('140728','140729','140737','140738','140739','140740','140741')
      AND SF2.F2_CLIENTE NOT IN ('002268','015699')
),
CTE_DEV_JMT AS (
    SELECT 'JMT' AS EMP,'DJ' AS TP,V.CODVEND,V.COD_SUPER AS SUPER,V.NOME_SUPER,V.VEND,SD1.D1_DOC AS NF,SD2.D2_PEDIDO AS PEDIDO,SD1.D1_ITEM AS ITEM,SD2.D2_COD AS CODPROD,SB1.B1_DESC AS PRODUTO,
           CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS DTEMISSAO,CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS VENCTO,CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS DTBAIXA,
           SA1.A1_COD AS CODCLIENTE,SA1.A1_ZZREDE AS REDE,SA1.A1_EST AS UF,SA1.A1_NOME AS CLIENTE,
           CASE SA1.A1_CODSEG WHEN '004000' THEN 'D' WHEN '001000' THEN 'V' WHEN '002000' THEN 'A' ELSE 'V' END AS CLASS_CLI,
           SD1.D1_TOTAL * -1 AS VLRBRUTO,SD1.D1_TOTAL * -1 AS VLRLIQ,SA1.A1_COMIS AS COMIS_CLI,V.COMIS_VEND,
           SB1.B1_COMIS AS COMIS_PROD,
           ((SD1.D1_TOTAL * SB1.B1_COMIS) / 100) * -1 AS TCOMISPROD
    FROM SD1020 SD1
    INNER JOIN SD2020 SD2 ON SD2.D2_DOC=SD1.D1_NFORI AND SD2.D2_SERIE=SD1.D1_SERIORI AND SD2.D2_ITEM=SD1.D1_ITEMORI AND SD2.D2_CLIENTE=SD1.D1_FORNECE AND SD2.D2_LOJA=SD1.D1_LOJA AND SD2.D_E_L_E_T_=''
    INNER JOIN SB1020 SB1 ON SB1.B1_COD=SD1.D1_COD AND SB1.D_E_L_E_T_=''
    INNER JOIN SA1020 SA1 ON SA1.A1_COD=SD1.D1_FORNECE AND SA1.A1_LOJA=SD1.D1_LOJA AND SA1.D_E_L_E_T_=''
    INNER JOIN SF2020 SF2 ON SF2.F2_DOC=SD2.D2_DOC AND SF2.F2_SERIE=SD2.D2_SERIE AND SF2.D_E_L_E_T_=''
    INNER JOIN CTE_VENDEDORES V ON V.CODVEND=SF2.F2_VEND1
    WHERE SD1.D_E_L_E_T_='' AND SD1.D1_TIPO='D' AND SUBSTRING(SD1.D1_CF,2,1)<>'9'
      AND CONVERT(VARCHAR(8), SD1.D1_DTDIGIT, 112) BETWEEN @DTBXI AND @DTBXF
      AND SD1.D1_DOC NOT IN ('140728','140729','140737','140738','140739','140740','140741')
      AND SF2.F2_CLIENTE NOT IN ('002268','015699')
),
CTE_DEV_3F AS (
    SELECT '3F' AS EMP,'D3' AS TP,V.CODVEND,V.COD_SUPER AS SUPER,V.NOME_SUPER,V.VEND,SD1.D1_DOC AS NF,SD2.D2_PEDIDO AS PEDIDO,SD1.D1_ITEM AS ITEM,SD2.D2_COD AS CODPROD,SB1.B1_DESC AS PRODUTO,
           CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS DTEMISSAO,CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS VENCTO,CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS DTBAIXA,
           SA1.A1_COD AS CODCLIENTE,SA1.A1_ZZREDE AS REDE,SA1.A1_EST AS UF,SA1.A1_NOME AS CLIENTE,
           CASE SA1.A1_CODSEG WHEN '004000' THEN 'D' WHEN '001000' THEN 'V' WHEN '002000' THEN 'A' ELSE 'V' END AS CLASS_CLI,
           SD1.D1_TOTAL * -1 AS VLRBRUTO,SD1.D1_TOTAL * -1 AS VLRLIQ,SA1.A1_COMIS AS COMIS_CLI,V.COMIS_VEND,
           SB1.B1_COMIS AS COMIS_PROD,
           ((SD1.D1_TOTAL * SB1.B1_COMIS) / 100) * -1 AS TCOMISPROD
    FROM SD1040 SD1
    INNER JOIN SD2040 SD2 ON SD2.D2_DOC=SD1.D1_NFORI AND SD2.D2_SERIE=SD1.D1_SERIORI AND SD2.D2_ITEM=SD1.D1_ITEMORI AND SD2.D2_CLIENTE=SD1.D1_FORNECE AND SD2.D2_LOJA=SD1.D1_LOJA AND SD2.D_E_L_E_T_=''
    INNER JOIN SB1040 SB1 ON SB1.B1_COD=SD1.D1_COD AND SB1.D_E_L_E_T_=''
    INNER JOIN SA1040 SA1 ON SA1.A1_COD=SD1.D1_FORNECE AND SA1.A1_LOJA=SD1.D1_LOJA AND SA1.D_E_L_E_T_=''
    INNER JOIN SF2040 SF2 ON SF2.F2_DOC=SD2.D2_DOC AND SF2.F2_SERIE=SD2.D2_SERIE AND SF2.D_E_L_E_T_=''
    INNER JOIN CTE_VENDEDORES V ON V.CODVEND=SF2.F2_VEND1
    WHERE SD1.D_E_L_E_T_='' AND SD1.D1_TIPO='D' AND SUBSTRING(SD1.D1_CF,2,1)<>'9'
      AND CONVERT(VARCHAR(8), SD1.D1_DTDIGIT, 112) BETWEEN @DTBXI AND @DTBXF
      AND SD1.D1_DOC NOT IN ('140728','140729','140737','140738','140739','140740','140741')
      AND SF2.F2_CLIENTE NOT IN ('002268','015699')
),
CTE_DEV_COS AS (
    SELECT 'COSMETICOS' AS EMP,'DC' AS TP,V.CODVEND,V.COD_SUPER AS SUPER,V.NOME_SUPER,V.VEND,SD1.D1_DOC AS NF,SD2.D2_PEDIDO AS PEDIDO,SD1.D1_ITEM AS ITEM,SD2.D2_COD AS CODPROD,SB1.B1_DESC AS PRODUTO,
           CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS DTEMISSAO,CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS VENCTO,CAST(CONVERT(VARCHAR(10),SD1.D1_DTDIGIT,103) AS DATE) AS DTBAIXA,
           SA1.A1_COD AS CODCLIENTE,SA1.A1_ZZREDE AS REDE,SA1.A1_EST AS UF,SA1.A1_NOME AS CLIENTE,
           CASE SA1.A1_CODSEG WHEN '004000' THEN 'D' WHEN '001000' THEN 'V' WHEN '002000' THEN 'A' ELSE 'V' END AS CLASS_CLI,
           SD1.D1_TOTAL * -1 AS VLRBRUTO,SD1.D1_TOTAL * -1 AS VLRLIQ,SA1.A1_COMIS AS COMIS_CLI,V.COMIS_VEND,
           SB1.B1_COMIS AS COMIS_PROD,
           ((SD1.D1_TOTAL * SB1.B1_COMIS) / 100) * -1 AS TCOMISPROD
    FROM SD1050 SD1
    INNER JOIN SD2050 SD2 ON SD2.D2_DOC=SD1.D1_NFORI AND SD2.D2_SERIE=SD1.D1_SERIORI AND SD2.D2_ITEM=SD1.D1_ITEMORI AND SD2.D2_CLIENTE=SD1.D1_FORNECE AND SD2.D2_LOJA=SD1.D1_LOJA AND SD2.D_E_L_E_T_=''
    INNER JOIN SB1050 SB1 ON SB1.B1_COD=SD1.D1_COD AND SB1.D_E_L_E_T_=''
    INNER JOIN SA1050 SA1 ON SA1.A1_COD=SD1.D1_FORNECE AND SA1.A1_LOJA=SD1.D1_LOJA AND SA1.D_E_L_E_T_=''
    INNER JOIN SF2050 SF2 ON SF2.F2_DOC=SD2.D2_DOC AND SF2.F2_SERIE=SD2.D2_SERIE AND SF2.D_E_L_E_T_=''
    INNER JOIN CTE_VENDEDORES V ON V.CODVEND=SF2.F2_VEND1
    WHERE SD1.D_E_L_E_T_='' AND SD1.D1_TIPO='D' AND SUBSTRING(SD1.D1_CF,2,1)<>'9'
      AND CONVERT(VARCHAR(8), SD1.D1_DTDIGIT, 112) BETWEEN @DTBXI AND @DTBXF
      AND SD1.D1_DOC NOT IN ('140728','140729','140737','140738','140739','140740','140741')
      AND SF2.F2_CLIENTE NOT IN ('002268','015699')
)
SELECT DISTINCT EMP,TP,CODVEND,SUPER,NOME_SUPER AS SUPER_NOME,VEND,NF,PEDIDO,ITEM,CODPROD,PRODUTO,DTEMISSAO,VENCTO,DTBAIXA,CODCLIENTE,REDE,UF,CLIENTE,CLASS_CLI,VLRBRUTO,VLRLIQ,COMIS_CLI,COMIS_VEND,COMIS_PROD,TCOMISPROD,MONTH(DTBAIXA) AS MES,YEAR(DTBAIXA) AS ANO
FROM (
    SELECT * FROM CTE_DEV_SOB
    UNION ALL SELECT * FROM CTE_DEV_JMT
    UNION ALL SELECT * FROM CTE_DEV_3F
    UNION ALL SELECT * FROM CTE_DEV_COS
) DEVOLUCOES
ORDER BY EMP,DTBAIXA,CODVEND,NF,ITEM;
"""


def _fetch_sql_resultsets(conn_str: str, sql: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    try:
        import pyodbc  # type: ignore
    except Exception as e:
        raise RuntimeError(f"driver_sql_indisponivel: {e}")

    conn = pyodbc.connect(conn_str, timeout=120)
    try:
        cur = conn.cursor()
        cur.execute(sql)
        resultsets: List[List[Dict[str, Any]]] = []
        while True:
            if cur.description:
                cols = [str(c[0] or "") for c in cur.description]
                rows = [dict(zip(cols, row)) for row in cur.fetchall()]
                resultsets.append(rows)
            if not cur.nextset():
                break
    finally:
        conn.close()

    if not resultsets:
        return [], []
    if len(resultsets) == 1:
        return resultsets[0], []
    return resultsets[0], resultsets[1]


def _conn_str_from_cfg(cfg: Dict[str, Any]) -> str:
    server = str(cfg.get("sql_server", "") or "").strip()
    if not server:
        return ""
    port = int(cfg.get("sql_port", 1433) or 1433)
    database = str(cfg.get("sql_database", "") or "").strip()
    user = str(cfg.get("sql_user", "") or "").strip()
    passwd = str(cfg.get("sql_pass", "") or "")
    encrypt = "yes" if int(cfg.get("sql_encrypt", 0) or 0) == 1 else "no"
    trust = "yes" if int(cfg.get("sql_trust_cert", 1) or 1) == 1 else "no"
    return (
        f"Driver={{ODBC Driver 18 for SQL Server}};Server={server},{port};Database={database};"
        f"UID={user};PWD={passwd};Encrypt={encrypt};TrustServerCertificate={trust};Connection Timeout=30;"
    )


def testar_conexao_sql(conn_str: Optional[str] = None, cfg: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    conn_final = str(conn_str or "").strip()
    if not conn_final:
        cfg_eff = dict(cfg or {})
        conn_final = _conn_str_from_cfg(cfg_eff)
    if not conn_final:
        raise ValueError("conn_string_sql_requerida")
    try:
        import pyodbc  # type: ignore
    except Exception as e:
        raise RuntimeError(f"driver_sql_indisponivel: {e}")

    conn = pyodbc.connect(conn_final, timeout=15)
    try:
        cur = conn.cursor()
        cur.execute("SELECT @@SERVERNAME AS servidor, DB_NAME() AS banco")
        row = cur.fetchone()
    finally:
        conn.close()
    return {
        "status": "ok",
        "servidor": str(row[0] or "") if row else "",
        "banco": str(row[1] or "") if row else "",
    }


def importar_vendas_query_banco(
    mes: int,
    ano: int,
    conn_str: Optional[str] = None,
    incluir_devolucoes: Optional[bool] = None,
) -> Dict[str, int]:
    init_schema()
    if not (1 <= int(mes or 0) <= 12):
        raise ValueError("mes_invalido")
    if not (2000 <= int(ano or 0) <= 2100):
        raise ValueError("ano_invalido")

    from ..database import models

    conn_final = str(conn_str or "").strip()
    cfg = {}
    if not conn_final:
        cfg = models.obter_configuracoes()
        conn_final = _conn_str_from_cfg(cfg)
    if not conn_final:
        raise ValueError("conn_string_sql_requerida")

    if incluir_devolucoes is None:
        incluir_devolucoes = True

    dtemissao_ini, dtemissao_fim, dtbaixa_ini, dtbaixa_fim = _periodo_query(int(mes), int(ano))
    sql = _query_comissoes_unificada(
        dtini=_to_yyyymmdd(dtemissao_ini),
        dtfim=_to_yyyymmdd(dtemissao_fim),
        dtbaixa_ini=_to_yyyymmdd(dtbaixa_ini),
        dtbaixa_fim=_to_yyyymmdd(dtbaixa_fim),
    )
    registros_db_raw, registros_dv_raw = _fetch_sql_resultsets(conn_final, sql)
    registros_db = [_norm_row_sql(r, int(mes), int(ano)) for r in registros_db_raw]
    registros_dv = [_norm_row_sql(r, int(mes), int(ano)) for r in registros_dv_raw] if incluir_devolucoes else []

    # Evita mistura de dados antigos com nova carga do mesmo periodo.
    # A limpeza por mes/ano remove tambem residuos legados sem apuracao_id consistente.
    _replace_period_data(int(mes), int(ano))

    assinatura = f"{mes:02d}/{ano}|{_to_yyyymmdd(dtemissao_ini)}|{_to_yyyymmdd(dtemissao_fim)}|{_to_yyyymmdd(dtbaixa_ini)}|{_to_yyyymmdd(dtbaixa_fim)}"
    hash_ref = hashlib.sha256(assinatura.encode("utf-8")).hexdigest()
    apuracao_id = criar_ou_substituir_apuracao_periodo(
        mes=int(mes),
        ano=int(ano),
        arquivo_nome=f"QUERY_SQL_{mes:02d}_{ano}",
        arquivo_hash=hash_ref,
        data_importacao=datetime.utcnow().isoformat(),
    )

    inserir_lancamentos(registros_db, "DB", apuracao_id=apuracao_id)
    inserir_lancamentos(registros_dv, "DS", apuracao_id=apuracao_id)
    _registrar_importacao(
        path=Path(f"QUERY_SQL_{mes:02d}_{ano}.sql"),
        file_hash=hash_ref,
        mes=int(mes),
        ano=int(ano),
        dbc=len(registros_db),
        devolucoes=len(registros_dv),
        apuracao_id=apuracao_id,
    )
    return {
        "dbc": len(registros_db),
        "devolucoes": len(registros_dv),
        "incluiu_devolucoes": bool(incluir_devolucoes),
        "mes": int(mes),
        "ano": int(ano),
        "apuracao_id": apuracao_id,
        "dtini": _to_yyyymmdd(dtemissao_ini),
        "dtfim": _to_yyyymmdd(dtemissao_fim),
        "dtbaixa_ini": _to_yyyymmdd(dtbaixa_ini),
        "dtbaixa_fim": _to_yyyymmdd(dtbaixa_fim),
    }

