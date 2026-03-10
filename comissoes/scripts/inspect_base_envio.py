from openpyxl import load_workbook
from pathlib import Path

p = Path(__file__).resolve().parents[2] / "artfatos" / "BASE ENVIO COMISSAO JANEIRO.xlsm"
wb = load_workbook(filename=str(p), data_only=True)
print(wb.sheetnames)
for name in wb.sheetnames[:5]:
    ws = wb[name]
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    print(name, [str(c).strip().lower() for c in row])
